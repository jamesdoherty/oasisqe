# coding=utf-8
"""
    Functionality for importing and exporting spreadsheets.
"""

from oasis.lib import Groups, Courses2, Exams, Users2

from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
from datetime import datetime


def exam_results_as_spreadsheet(course_id, group, exam_id):
    """ Export the assessment results as a XLSX spreadsheet """

    course = Courses2.get_course(course_id)
    exam = Exams.get_exam_struct(exam_id, course_id)

    uids = set([])
    totals = {}

    results = Exams.get_marks(group, exam_id)
    for user_id in results:
        uids.add(user_id)
        if not user_id in totals:
            totals[user_id] = 0.0
        for qt, val in results[user_id].iteritems():
            totals[user_id] += val['score']

    questions = Exams.get_qts_list(exam_id)
    users = {}
    for uid in uids:
        users[uid] = Users2.get_user(uid)

    when = datetime.now().strftime("%H:%m, %a %d %b %Y"),

    wb = Workbook()

    ws = wb.get_active_sheet()

    ws.title = "Results"

    ws.cell(row=1, column=0).value = course['name']
    ws.cell(row=1, column=1).value = course['title']
    ws.cell(row=2, column=0).value = "Assessment:"
    ws.cell(row=2, column=1).value = exam['title']
    ws.cell(row=3, column=0).value = "Group:"
    ws.cell(row=3, column=1).value = group.name

    col = 5
    qcount = 1
    for _ in questions:
        ws.cell(row=4, column=col).value = "Q%s" % qcount
        qcount += 1
        col += 1

    ws.cell(row=4, column=col).value = "Total"

    row = 5
    sortusers = users.keys()
    sortusers.sort(key=lambda us: users[us]['familyname'])

    for user_id in sortusers:
        result = results[user_id]
        ws.cell(row=row, column=0).value = users[user_id]['uname']
        ws.cell(row=row, column=1).value = users[user_id]['student_id']
        ws.cell(row=row, column=2).value = users[user_id]['familyname']
        ws.cell(row=row, column=3).value = users[user_id]['givenname']
        ws.cell(row=row, column=4).value = users[user_id]['email']
        col = 5

        for pos in questions:
            for qt in pos:
                if qt['id'] in result:
                    ws.cell(row=row, column=col).value = result[qt['id']]['score']
                    col += 1

        ws.cell(row=row, column=col).value = totals[user_id]
        row += 1
  #  ws.cell(row=2,column=3).value = "NCEA4"

  # for q in range(11,56):    # questions
  #      ws.cell(row=1,column=x).value="%s" % (q,)
  #      ws.cell(row=3, column=x).value=correct["q%s"%(q,)]

  #  for row in [1,2,3]:
  #      for col in range(0,121):
  #          ws.cell(row=row, column=col).style.font.bold=True

    return save_virtual_workbook(wb)